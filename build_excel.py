"""
Comprobante de Egreso — Hacienda La Petrolea
Excel version: two copies on one Letter-size sheet (8.5×11in)
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
import os

wb = Workbook()
ws = wb.active
ws.title = "Comprobante"

# ── Page setup (Letter, portrait, narrow margins) ──────────────────────────
ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
ws.page_setup.orientation = "portrait"
ws.page_setup.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.page_margins.left   = 0.35
ws.page_margins.right  = 0.35
ws.page_margins.top    = 0.30
ws.page_margins.bottom = 0.30
ws.page_margins.header = 0
ws.page_margins.footer = 0
ws.sheet_view.showGridLines = False

# ── Helpers ────────────────────────────────────────────────────────────────
def thin():  return Side(style='thin',   color='000000')
def thick(): return Side(style='medium', color='000000')
def hair():  return Side(style='hair',   color='999999')
def dash():  return Side(style='dashed', color='AAAAAA')
def none():  return Side(style=None)

def box(top=None, bottom=None, left=None, right=None):
    return Border(top=top or none(), bottom=bottom or none(),
                  left=left or none(), right=right or none())

def full_box(style='thin'):
    s = Side(style=style, color='000000')
    return Border(top=s, bottom=s, left=s, right=s)

def cell(row, col, value='', bold=False, size=9, align='left', valign='center',
         wrap=False, fill_color=None, border=None, num_format=None,
         italic=False, color='000000', uppercase=False):
    c = ws.cell(row=row, column=col)
    if uppercase and isinstance(value, str):
        value = value.upper()
    c.value = value
    c.font = Font(name='Georgia', bold=bold, size=size, color=color,
                  italic=italic)
    c.alignment = Alignment(horizontal=align, vertical=valign,
                            wrap_text=wrap)
    if fill_color:
        c.fill = PatternFill('solid', fgColor=fill_color)
    if border:
        c.border = border
    if num_format:
        c.number_format = num_format
    return c

def merge(r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1,
                   end_row=r2, end_column=c2)

def set_row_height(row, pts):
    ws.row_dimensions[row].height = pts

def set_col_width(col, chars):
    ws.column_dimensions[get_column_letter(col)].width = chars

# ── Column layout (A–L = 12 columns across 7.80in usable) ─────────────────
#  A  B  C  D  E  F  G  H  I  J  K  L
#  logo  |      company info       |  doc block
# Field rows span A–L with varying merges

col_widths = [
    1.2,   # A  left padding
    5.5,   # B  logo / left col
    6.0,   # C
    6.0,   # D
    6.0,   # E
    6.0,   # F
    6.0,   # G
    6.0,   # H
    6.0,   # I
    5.5,   # J
    5.5,   # K
    8.0,   # L  right doc block
]
for i, w in enumerate(col_widths, 1):
    set_col_width(i, w)

# ── Build one copy starting at `start_row` ────────────────────────────────
def build_copy(start_row, copy_label):
    r = start_row
    HDR_FILL  = 'F2F2F2'
    SEC_FILL  = 'FFFFFF'
    DARK_FILL = 'E8E8E8'

    # ══════════════════════════════════════════════
    # HEADER — rows r to r+4
    # ══════════════════════════════════════════════
    # Top border row
    for c in range(2, 13):
        ws.cell(row=r, column=c).border = box(top=thick())
    set_row_height(r, 3)
    r += 1  # r+1 = header start

    hdr_rows = 5
    hdr_end  = r + hdr_rows - 1

    # Left border on all header rows
    for row in range(r, hdr_end + 1):
        for c in range(2, 13):
            cur = ws.cell(row=row, column=c).border
            ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor=HDR_FILL)

    # Row 1: Company name + doc title
    set_row_height(r, 20)
    merge(r, 2, r, 9)
    c_name = ws.cell(row=r, column=2)
    c_name.value = 'HACIENDA LA PETROLEA'
    c_name.font = Font(name='Georgia', bold=True, size=14, color='000000')
    c_name.alignment = Alignment(horizontal='center', vertical='center')
    c_name.fill = PatternFill('solid', fgColor=HDR_FILL)
    c_name.border = box(left=thick())

    merge(r, 10, r+1, 12)
    dt = ws.cell(row=r, column=10)
    dt.value = 'Comprobante de Egreso'
    dt.font = Font(name='Georgia', bold=True, size=11, color='000000')
    dt.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
    dt.fill = PatternFill('solid', fgColor=HDR_FILL)
    dt.border = box(right=thick(), left=thin())

    # Row 2: Est line
    r += 1
    set_row_height(r, 11)
    merge(r, 2, r, 9)
    est = ws.cell(row=r, column=2)
    est.value = 'Est. 1943  •  La Dorada, Caldas, Colombia'
    est.font = Font(name='Arial', size=7, color='666666')
    est.alignment = Alignment(horizontal='center', vertical='center')
    est.fill = PatternFill('solid', fgColor=HDR_FILL)
    est.border = box(left=thick())
    # col 10-12 merged above

    # Row 3: Separator + Número label
    r += 1
    set_row_height(r, 3)
    merge(r, 2, r, 9)
    sep = ws.cell(row=r, column=2)
    sep.border = box(left=thick(), bottom=hair())
    sep.fill = PatternFill('solid', fgColor=HDR_FILL)

    merge(r, 10, r, 12)
    num_lbl = ws.cell(row=r, column=10)
    num_lbl.value = 'NÚMERO'
    num_lbl.font = Font(name='Arial', size=6, color='666666', bold=True)
    num_lbl.alignment = Alignment(horizontal='center', vertical='center')
    num_lbl.fill = PatternFill('solid', fgColor=DARK_FILL)
    num_lbl.border = box(top=thin(), bottom=thin(), left=thin(), right=thick())

    # Row 4: Propietaria info + Número value
    r += 1
    set_row_height(r, 14)
    merge(r, 2, r, 9)
    info = ws.cell(row=r, column=2)
    info.value = 'Propietaria: Patricia Platts de Hurtado    NIT: 680-745-545    Actividad: Ganadería & Agricultura'
    info.font = Font(name='Georgia', size=8, color='333333')
    info.alignment = Alignment(horizontal='left', vertical='center')
    info.fill = PatternFill('solid', fgColor=HDR_FILL)
    info.border = box(left=thick())

    merge(r, 10, r, 12)
    num_val = ws.cell(row=r, column=10)
    num_val.value = ''
    num_val.font = Font(name='Georgia', bold=True, size=13)
    num_val.alignment = Alignment(horizontal='center', vertical='center')
    num_val.fill = PatternFill('solid', fgColor='FFFFFF')
    num_val.border = box(bottom=dash(), left=thin(), right=thick())

    # Row 5: Copy tag
    r += 1
    set_row_height(r, 11)
    merge(r, 2, r, 9)
    dummy = ws.cell(row=r, column=2)
    dummy.fill = PatternFill('solid', fgColor=HDR_FILL)
    dummy.border = box(left=thick(), bottom=thick())
    # Apply bottom border to full row
    for c in range(3, 10):
        ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=HDR_FILL)
        ws.cell(row=r, column=c).border = box(bottom=thick())

    merge(r, 10, r, 12)
    copy_tag = ws.cell(row=r, column=10)
    copy_tag.value = copy_label
    copy_tag.font = Font(name='Arial', size=7, bold=True, color='333333')
    copy_tag.alignment = Alignment(horizontal='center', vertical='center')
    copy_tag.fill = PatternFill('solid', fgColor=DARK_FILL)
    copy_tag.border = box(top=thin(), bottom=thick(), left=thin(), right=thick())

    r += 1  # ── end of header ──

    # ══════════════════════════════════════════════
    # SECTION: DATOS DEL PAGO
    # ══════════════════════════════════════════════
    set_row_height(r, 11)
    merge(r, 2, r, 12)
    sec1 = ws.cell(row=r, column=2)
    sec1.value = 'DATOS DEL PAGO'
    sec1.font = Font(name='Arial', size=7, bold=True, color='666666')
    sec1.alignment = Alignment(horizontal='left', vertical='center')
    sec1.border = box(left=thick(), right=thick(), bottom=hair())
    r += 1

    # Fecha / Pagado a / Cédula-NIT
    set_row_height(r, 7)
    for lbl, cols in [('FECHA', (2,3)), ('PAGADO A', (4,8)), ('CÉDULA / NIT', (9,12))]:
        c = ws.cell(row=r, column=cols[0])
        merge(r, cols[0], r, cols[1])
        c.value = lbl
        c.font = Font(name='Arial', size=6, bold=True, color='888888')
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.border = box(left=thick() if cols[0]==2 else none(),
                       right=thick() if cols[1]==12 else none())
    r += 1

    set_row_height(r, 16)
    for cols in [(2,3), (4,8), (9,12)]:
        c = ws.cell(row=r, column=cols[0])
        merge(r, cols[0], r, cols[1])
        c.value = ''
        c.border = box(
            bottom=thin(),
            left=thick() if cols[0]==2 else none(),
            right=thick() if cols[1]==12 else none()
        )
    r += 1

    # Cargo
    set_row_height(r, 7)
    merge(r, 2, r, 12)
    c = ws.cell(row=r, column=2)
    c.value = 'CARGO / RELACIÓN CON LA EMPRESA'
    c.font = Font(name='Arial', size=6, bold=True, color='888888')
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = box(left=thick(), right=thick())
    r += 1

    set_row_height(r, 16)
    merge(r, 2, r, 12)
    c = ws.cell(row=r, column=2)
    c.border = box(bottom=thin(), left=thick(), right=thick())
    r += 1

    # ── Forma de pago ──
    set_row_height(r, 11)
    merge(r, 2, r, 12)
    c = ws.cell(row=r, column=2)
    c.value = 'FORMA DE PAGO'
    c.font = Font(name='Arial', size=6, bold=True, color='888888')
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = box(left=thick(), right=thick())
    r += 1

    set_row_height(r, 16)
    # Checkboxes row: ☐ Depósito Directo  ☐ Efectivo  ☐ Cheque N°. _____
    merge(r, 2, r, 4)
    c = ws.cell(row=r, column=2)
    c.value = '☐  Depósito Directo'
    c.font = Font(name='Arial', size=9)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = box(left=thick())

    merge(r, 5, r, 7)
    c = ws.cell(row=r, column=5)
    c.value = '☐  Efectivo'
    c.font = Font(name='Arial', size=9)
    c.alignment = Alignment(horizontal='left', vertical='center')

    merge(r, 8, r, 12)
    c = ws.cell(row=r, column=8)
    c.value = '☐  Cheque N°. ___________________'
    c.font = Font(name='Arial', size=9)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = box(right=thick())
    r += 1

    # ── Valor total ──
    set_row_height(r, 22)
    merge(r, 2, r, 5)
    c = ws.cell(row=r, column=2)
    c.value = 'VALOR TOTAL — COP $'
    c.font = Font(name='Arial', size=7, bold=True, color='666666')
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.fill = PatternFill('solid', fgColor='F8F8F8')
    c.border = box(top=thin(), bottom=thin(), left=thick())

    merge(r, 6, r, 12)
    c = ws.cell(row=r, column=6)
    c.value = ''
    c.font = Font(name='Georgia', bold=True, size=14)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = box(top=thin(), bottom=thin(), left=none(), right=thick())
    r += 1

    # ── La suma de ──
    set_row_height(r, 14)
    merge(r, 2, r, 3)
    c = ws.cell(row=r, column=2)
    c.value = 'LA SUMA DE'
    c.font = Font(name='Arial', size=6, bold=True, color='888888')
    c.alignment = Alignment(horizontal='left', vertical='bottom')
    c.border = box(left=thick())

    merge(r, 4, r, 12)
    c = ws.cell(row=r, column=4)
    c.border = box(bottom=thin(), right=thick())
    r += 1

    # ══════════════════════════════════════════════
    # SECTION: DETALLE DEL CONCEPTO
    # ══════════════════════════════════════════════
    set_row_height(r, 11)
    merge(r, 2, r, 12)
    c = ws.cell(row=r, column=2)
    c.value = 'DETALLE DEL CONCEPTO'
    c.font = Font(name='Arial', size=7, bold=True, color='666666')
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = box(left=thick(), right=thick(), top=thin())
    r += 1

    # Table header
    set_row_height(r, 13)
    merge(r, 2, r, 8)
    c = ws.cell(row=r, column=2)
    c.value = 'DESCRIPCIÓN'
    c.font = Font(name='Arial', size=7, bold=True, color='666666')
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.fill = PatternFill('solid', fgColor=DARK_FILL)
    c.border = box(top=thick(), bottom=thick(), left=thick())

    merge(r, 9, r, 10)
    c = ws.cell(row=r, column=9)
    c.value = 'CANT.'
    c.font = Font(name='Arial', size=7, bold=True, color='666666')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill('solid', fgColor=DARK_FILL)
    c.border = box(top=thick(), bottom=thick(), left=thin(), right=thin())

    merge(r, 11, r, 12)
    c = ws.cell(row=r, column=11)
    c.value = 'VALOR'
    c.font = Font(name='Arial', size=7, bold=True, color='666666')
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.fill = PatternFill('solid', fgColor=DARK_FILL)
    c.border = box(top=thick(), bottom=thick(), right=thick())
    r += 1

    # 3 blank rows
    for _ in range(3):
        set_row_height(r, 16)
        merge(r, 2, r, 8)
        ws.cell(row=r, column=2).border = box(bottom=hair(), left=thick())
        merge(r, 9, r, 10)
        ws.cell(row=r, column=9).border = box(bottom=hair(), left=thin(), right=thin())
        merge(r, 11, r, 12)
        ws.cell(row=r, column=11).border = box(bottom=hair(), right=thick())
        r += 1

    # TOTAL row
    set_row_height(r, 16)
    merge(r, 2, r, 10)
    c = ws.cell(row=r, column=2)
    c.value = 'TOTAL'
    c.font = Font(name='Georgia', bold=True, size=9)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.fill = PatternFill('solid', fgColor=DARK_FILL)
    c.border = box(top=thick(), bottom=thick(), left=thick())

    merge(r, 11, r, 12)
    c = ws.cell(row=r, column=11)
    c.value = '$   ___________________'
    c.font = Font(name='Georgia', bold=True, size=9)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.fill = PatternFill('solid', fgColor=DARK_FILL)
    c.border = box(top=thick(), bottom=thick(), right=thick())
    r += 1

    # ══════════════════════════════════════════════
    # SECTION: AUTORIZACIONES Y FIRMAS
    # ══════════════════════════════════════════════
    set_row_height(r, 10)
    merge(r, 2, r, 12)
    c = ws.cell(row=r, column=2)
    c.value = 'AUTORIZACIONES Y FIRMAS'
    c.font = Font(name='Arial', size=7, bold=True, color='666666')
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = box(top=thin(), left=thick(), right=thick())
    r += 1

    # Sig column headers
    set_row_height(r, 11)
    sig_cols = [
        ('HUELLA',                  2,  3),
        ('FIRMA DEL RECEPTOR',      4,  7),
        ('CÉDULA',                  8,  9),
        ('RECIBÍ CONFORME — FECHA', 10, 12),
    ]
    for lbl, c1, c2 in sig_cols:
        merge(r, c1, r, c2)
        c = ws.cell(row=r, column=c1)
        c.value = lbl
        c.font = Font(name='Arial', size=6, bold=True, color='666666')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill = PatternFill('solid', fgColor=DARK_FILL)
        c.border = box(
            top=thick(), bottom=thin(),
            left=thick() if c1 == 2 else thin(),
            right=thick() if c2 == 12 else none()
        )
    r += 1

    # Sig area rows (3 rows tall for space)
    for row_off in range(3):
        set_row_height(r, 18)
        for c1, c2 in [(2,3),(4,7),(8,9),(10,12)]:
            merge(r, c1, r, c2) if row_off == 0 else None
            c = ws.cell(row=r, column=c1)
            is_last_row = (row_off == 2)
            c.border = box(
                bottom=thick() if is_last_row else none(),
                left=thick() if c1 == 2 else thin(),
                right=thick() if c2 == 12 else none()
            )
            # Add oval hint in fingerprint cell on middle row
            if c1 == 2 and row_off == 1:
                c.value = '( )'
                c.font = Font(name='Arial', size=18, color='CCCCCC')
                c.alignment = Alignment(horizontal='center', vertical='center')
        r += 1

    # ══════════════════════════════════════════════
    # FOOTER
    # ══════════════════════════════════════════════
    set_row_height(r, 13)
    merge(r, 2, r, 9)
    c = ws.cell(row=r, column=2)
    c.value = 'Autopista Honda Km 8  •  La Dorada, Caldas    |    petrolea11@hotmail.com    |    Tel. 311-871-7493'
    c.font = Font(name='Arial', size=7, color='555555')
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.fill = PatternFill('solid', fgColor=DARK_FILL)
    c.border = box(top=thick(), bottom=thick(), left=thick())
    for c2 in range(3, 10):
        ws.cell(row=r, column=c2).fill = PatternFill('solid', fgColor=DARK_FILL)
        ws.cell(row=r, column=c2).border = box(top=thick(), bottom=thick())

    merge(r, 10, r, 12)
    c = ws.cell(row=r, column=10)
    c.value = 'CONSECUTIVO: ___________'
    c.font = Font(name='Arial', size=8, bold=True)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill('solid', fgColor='FFFFFF')
    c.border = full_box('medium')
    r += 1

    # Bottom border
    for col in range(2, 13):
        ws.cell(row=r, column=col).border = box(top=thick())
    set_row_height(r, 2)
    r += 1

    return r


# ── Build Copy 1 ───────────────────────────────────────────────────────────
next_row = build_copy(1, 'COPIA EMPRESA')

# ── Cut line ──────────────────────────────────────────────────────────────
cut_row = next_row
set_row_height(cut_row, 14)
merge(cut_row, 2, cut_row, 12)
c = ws.cell(row=cut_row, column=2)
c.value = '✂   CORTE AQUÍ — COPIA PARA EL RECEPTOR   ✂'
c.font = Font(name='Arial', size=7, color='888888', italic=True)
c.alignment = Alignment(horizontal='center', vertical='center')
c.border = box(top=dash(), bottom=dash())
next_row = cut_row + 1

# ── Build Copy 2 ───────────────────────────────────────────────────────────
build_copy(next_row, 'COPIA RECEPTOR')

# ── Save ───────────────────────────────────────────────────────────────────
out = '/home/andres/.openclaw/workspace/petrolea-receipt/Comprobante-Egreso-Petrolea.xlsx'
wb.save(out)
print(f'Saved: {out}')
